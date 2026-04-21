"""Excel generation and reading — produces review file and reads back validated data."""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
NEW_ROW_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def generate_excel(
    filepath: str | Path,
    target_date: str,
    existing_records: list[dict],
    proposed_entries: list[dict],
    tasks_lookup: dict[int, dict],
) -> Path:
    """Create an Excel file with two sheets: Existing time + Proposed new entries.

    existing_records: flat list of { taskId, time, comment?, taskName?, projectName?, customerName? }
    proposed_entries: list of { taskId, taskName, minutes, comment }
    tasks_lookup: { taskId: { name, projectName, customerName } }
    """
    filepath = Path(filepath)
    wb = Workbook()

    # ── Sheet 1: Existing time-track ──────────────────────────────
    ws_existing = wb.active
    ws_existing.title = "Existing"
    headers_ex = ["Task ID", "Customer", "Project", "Task", "Time (min)", "Time (h)", "Comment"]
    ws_existing.append(headers_ex)
    _style_header(ws_existing, len(headers_ex))

    total_existing = 0
    for rec in existing_records:
        tid = rec.get("taskId", "")
        t = tasks_lookup.get(tid, {})
        minutes = rec.get("time", 0)
        total_existing += minutes
        ws_existing.append([
            tid,
            rec.get("customerName", t.get("customerName", "")),
            rec.get("projectName", t.get("projectName", "")),
            rec.get("taskName", t.get("name", "")),
            minutes,
            round(minutes / 60, 2),
            rec.get("comment", ""),
        ])

    # Totals row
    row_ex = ws_existing.max_row + 1
    ws_existing.cell(row=row_ex, column=4, value="TOTAL").font = Font(bold=True)
    ws_existing.cell(row=row_ex, column=5, value=total_existing).font = Font(bold=True)
    ws_existing.cell(row=row_ex, column=6, value=round(total_existing / 60, 2)).font = Font(bold=True)

    # ── Sheet 2: Proposed entries (editable) ──────────────────────
    ws_proposed = wb.create_sheet("Proposed")
    headers_pr = ["Task ID", "Customer", "Project", "Task", "Minutes", "Hours", "Comment", "Include? (Y/N)"]
    ws_proposed.append(headers_pr)
    _style_header(ws_proposed, len(headers_pr))

    total_proposed = 0
    for entry in proposed_entries:
        tid = entry.get("taskId", "")
        t = tasks_lookup.get(tid, {})
        minutes = entry.get("minutes", 0)
        total_proposed += minutes
        row_data = [
            tid,
            t.get("customerName", ""),
            t.get("projectName", ""),
            entry.get("taskName", t.get("name", "")),
            minutes,
            round(minutes / 60, 2),
            entry.get("comment", ""),
            "Y",
        ]
        ws_proposed.append(row_data)
        # Highlight new rows in green
        for col in range(1, len(row_data) + 1):
            cell = ws_proposed.cell(row=ws_proposed.max_row, column=col)
            cell.fill = NEW_ROW_FILL
            cell.border = THIN_BORDER

    # Totals row
    row_pr = ws_proposed.max_row + 1
    ws_proposed.cell(row=row_pr, column=4, value="TOTAL").font = Font(bold=True)
    ws_proposed.cell(row=row_pr, column=5, value=total_proposed).font = Font(bold=True)
    ws_proposed.cell(row=row_pr, column=6, value=round(total_proposed / 60, 2)).font = Font(bold=True)

    # Auto-size columns (rough)
    for ws in [ws_existing, ws_proposed]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(filepath)
    return filepath


def read_proposed_excel(filepath: str | Path) -> list[dict]:
    """Read back the 'Proposed' sheet after user review.

    Returns only rows where Include? == 'Y'.
    """
    filepath = Path(filepath)
    wb = load_workbook(filepath, read_only=True)
    ws = wb["Proposed"]

    entries = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if row[0] is None or str(row[0]).strip() == "":
            continue  # skip empty / totals row
        if str(row[0]).strip().upper() == "TOTAL":
            continue

        include = str(row[7]).strip().upper() if len(row) > 7 and row[7] else "Y"
        if include != "Y":
            continue

        entries.append({
            "taskId": int(row[0]),
            "minutes": int(row[4]) if row[4] else 0,
            "comment": str(row[6]) if row[6] else "",
        })

    wb.close()
    return entries
